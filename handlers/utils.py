import sqlite3
import pytesseract
from PIL import Image
import io
import re
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np
import logging

from config import DB_PATH
from .constants import *

logger = logging.getLogger(__name__)


def init_db():
    """Инициализация базы данных"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS daily_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL UNIQUE,
            day_of_week TEXT,
            sales REAL NOT NULL,
            checks INTEGER NOT NULL,
            cost_price REAL NOT NULL,
            expenses REAL NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sales_target REAL,
            net_profit_target REAL,
            expenses_target REAL,
            month TEXT,
            year INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    conn.commit()
    conn.close()


def add_data_to_db(date, day_of_week, sales, checks, cost_price, expenses):
    """Добавление данных в БД"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT OR REPLACE INTO daily_data 
            (date, day_of_week, sales, checks, cost_price, expenses)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (date, day_of_week, sales, checks, cost_price, expenses))
        conn.commit()
        success = True
    except Exception as e:
        logger.error(f"Error adding data: {e}")
        success = False
    finally:
        conn.close()
    
    return success


def get_all_data():
    """Получение всех данных"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT date, day_of_week, sales, checks, cost_price, expenses 
        FROM daily_data 
        ORDER BY date
    """)
    rows = cursor.fetchall()
    conn.close()
    return rows


def delete_data(date):
    """Удаление данных за конкретную дату"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM daily_data WHERE date = ?", (date,))
    conn.commit()
    conn.close()


def set_targets_db(sales_target, net_profit_target, expenses_target, month, year):
    """Установка целей"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    cursor.execute("""
        DELETE FROM targets 
        WHERE month = ? AND year = ?
    """, (month, year))
    
    cursor.execute("""
        INSERT INTO targets (sales_target, net_profit_target, expenses_target, month, year)
        VALUES (?, ?, ?, ?, ?)
    """, (sales_target, net_profit_target, expenses_target, month, year))
    
    conn.commit()
    conn.close()


def get_targets_db(month, year):
    """Получение целей"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT sales_target, net_profit_target, expenses_target
        FROM targets
        WHERE month = ? AND year = ?
        ORDER BY created_at DESC
        LIMIT 1
    """, (month, year))
    result = cursor.fetchone()
    conn.close()
    return result


def parse_date(date_str):
    """Парсинг даты и определение дня недели"""
    days_ru = ['пн', 'вт', 'ср', 'чт', 'пт', 'сб', 'вс']
    
    try:
        for fmt in ['%d.%m', '%d.%m.%Y', '%Y-%m-%d']:
            try:
                dt = datetime.strptime(date_str, fmt)
                if fmt == '%d.%m':
                    dt = dt.replace(year=datetime.now().year)
                day_of_week = days_ru[dt.weekday()]
                return dt.strftime('%d.%m'), day_of_week
            except ValueError:
                continue
        return None, None
    except:
        return None, None


def extract_from_photo(photo_bytes):
    """Извлечение данных из фото через OCR"""
    try:
        image = Image.open(io.BytesIO(photo_bytes))
        text = pytesseract.image_to_string(image, lang='rus')
        
        numbers = re.findall(r'\d+(?:[.,]\d+)?', text)
        
        values = []
        for num in numbers:
            try:
                values.append(float(num.replace(',', '.')))
            except:
                pass
        
        return values[:5] if len(values) >= 4 else None
        
    except Exception as e:
        logger.error(f"OCR Error: {e}")
        return None


def generate_excel_report():
    """Генерация Excel-отчёта в формате как на скриншоте"""
    data = get_all_data()
    targets_data = get_targets_db(
        datetime.now().strftime('%B'), 
        datetime.now().year
    )
    
    if not data:
        return None
    
    # Цели (или значения по умолчанию)
    if targets_data:
        sales_target = targets_data[0]
        profit_target = targets_data[1]
        expenses_target = targets_data[2]
    else:
        sales_target = 500000
        profit_target = 200000
        expenses_target = 300000
    
    # Создаем workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ПНЛ"
    
    # ============================================
    # СТИЛИ
    # ============================================
    header_font = Font(bold=True, size=11, color="FFFFFF")
    title_font = Font(bold=True, size=12)
    normal_font = Font(size=10)
    
    # Заливки
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    gray_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # ============================================
    # ШАПКА ОТЧЁТА (строки 1-2)
    # ============================================
    # Строка 1: Заголовок
    ws.merge_cells('C1:D1')
    ws['C1'] = "Отчёт:"
    ws['C1'].font = title_font
    ws['C1'].alignment = center_align
    
    ws['E1'] = "ПНЛ"
    ws['E1'].font = title_font
    ws['E1'].alignment = center_align
    
    ws['F1'] = "Май"
    ws['F1'].font = title_font
    ws['F1'].alignment = center_align
    
    ws.merge_cells('G1:H1')
    ws['G1'] = "2025"
    ws['G1'].font = title_font
    ws['G1'].alignment = center_align
    
    ws.merge_cells('I1:K1')
    ws['I1'] = "Название локации"
    ws['I1'].font = title_font
    ws['I1'].alignment = center_align
    
    # Строка 2: Заголовки колонок
    headers = [
        "Д/н",      # A2
        "Дата",     # B2
        "Продажи",  # C2
        "Чеки",     # D2
        "Ср чек",   # E2
        "Себестоимость",  # F2 (сумма)
        "",         # G2 (процент)
        "Себестоимость",  # H2 (сумма) - дублируется на скриншоте
        "",         # I2 (процент)
        "Валовая прибыль",  # J2 (сумма)
        "",         # K2 (процент)
        "Расходы",  # L2 (сумма)
        "",         # M2 (процент)
        "Чистая прибыль",  # N2 (сумма)
        ""          # O2 (процент)
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # ============================================
    # РАСЧЁТЫ
    # ============================================
    total_sales = sum(row[2] for row in data)
    total_checks = sum(row[3] for row in data)
    total_cost = sum(row[4] for row in data)
    total_expenses = sum(row[5] for row in data)
    total_gross_profit = total_sales - total_cost
    total_net_profit = total_gross_profit - total_expenses
    
    days_count = len(data)
    
    # Прогноз на месяц (30 дней)
    avg_daily_sales = total_sales / days_count if days_count > 0 else 0
    avg_daily_profit = total_net_profit / days_count if days_count > 0 else 0
    forecast_sales = avg_daily_sales * 30
    forecast_profit = avg_daily_profit * 30
    
    # Отклонения
    sales_deviation = total_sales - sales_target
    profit_deviation = total_net_profit - profit_target
    expenses_deviation = total_expenses - expenses_target
    
    # ============================================
    # СТРОКА 3: ОТКЛОНЕНИЕ
    # ============================================
    row = 3
    ws.cell(row=row, column=1, value="Отклонение").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sales_deviation).font = Font(color="FF0000" if sales_deviation < 0 else "008000")
    ws.cell(row=row, column=3, value=total_checks - int(sales_target / avg_daily_sales if avg_daily_sales > 0 else 0))
    ws.cell(row=row, column=4, value=0)
    ws.cell(row=row, column=5, value=total_cost)
    ws.cell(row=row, column=6, value=f"{total_cost/total_sales*100:.0f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=7, value=total_expenses)
    ws.cell(row=row, column=8, value=f"{total_expenses/total_sales*100:.0f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=9, value=total_gross_profit)
    ws.cell(row=row, column=10, value=f"{total_gross_profit/total_sales*100:.0f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=11, value=total_expenses)
    ws.cell(row=row, column=12, value=f"{total_expenses/total_sales*100:.0f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=13, value=total_net_profit).font = Font(color="FF0000" if total_net_profit < 0 else "008000")
    ws.cell(row=row, column=14, value=f"{total_net_profit/total_sales*100:.0f}%" if total_sales > 0 else "0%")
    
    # ============================================
    # СТРОКА 4: ЦЕЛЬ
    # ============================================
    row = 4
    ws.cell(row=row, column=1, value="Цель").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sales_target)
    ws.cell(row=row, column=3, value=int(sales_target / avg_daily_sales if avg_daily_sales > 0 else 0))
    ws.cell(row=row, column=4, value=0)
    ws.cell(row=row, column=5, value=sales_target * 0.37)
    ws.cell(row=row, column=6, value="37%")
    ws.cell(row=row, column=7, value=expenses_target)
    ws.cell(row=row, column=8, value="100%")
    ws.cell(row=row, column=9, value=sales_target * 0.60)
    ws.cell(row=row, column=10, value="60%")
    ws.cell(row=row, column=11, value=expenses_target)
    ws.cell(row=row, column=12, value="100%")
    ws.cell(row=row, column=13, value=profit_target)
    ws.cell(row=row, column=14, value=f"{profit_target/sales_target*100:.0f}%")
    
    # ============================================
    # СТРОКА 5: ПРОГНОЗ
    # ============================================
    row = 5
    ws.cell(row=row, column=1, value="Прогноз").font = Font(bold=True)
    ws.cell(row=row, column=2, value=forecast_sales)
    ws.cell(row=row, column=3, value=int(forecast_sales / avg_daily_sales if avg_daily_sales > 0 else 0))
    ws.cell(row=row, column=4, value=0)
    ws.cell(row=row, column=5, value=forecast_sales * 0.39)
    ws.cell(row=row, column=6, value="39%")
    ws.cell(row=row, column=7, value=expenses_target)
    ws.cell(row=row, column=8, value="100%")
    ws.cell(row=row, column=9, value=forecast_sales * 0.58)
    ws.cell(row=row, column=10, value="58%")
    ws.cell(row=row, column=11, value=expenses_target)
    ws.cell(row=row, column=12, value="100%")
    ws.cell(row=row, column=13, value=forecast_profit)
    ws.cell(row=row, column=14, value=f"{forecast_profit/forecast_sales*100:.0f}%" if forecast_sales > 0 else "0%")
    
    # ============================================
    # СТРОКА 6: НАКОПИТЕЛЬНО (итоги)
    # ============================================
    row = 6
    ws.cell(row=row, column=1, value="Накопительно").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_sales)
    ws.cell(row=row, column=3, value=total_checks)
    ws.cell(row=row, column=4, value=0)
    ws.cell(row=row, column=5, value=total_cost)
    ws.cell(row=row, column=6, value=f"{total_cost/total_sales*100:.1f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=7, value=total_expenses)
    ws.cell(row=row, column=8, value=f"{total_expenses/total_sales*100:.1f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=9, value=total_gross_profit)
    ws.cell(row=row, column=10, value=f"{total_gross_profit/total_sales*100:.1f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=11, value=total_expenses)
    ws.cell(row=row, column=12, value=f"{total_expenses/total_sales*100:.1f}%" if total_sales > 0 else "0%")
    ws.cell(row=row, column=13, value=total_net_profit)
    ws.cell(row=row, column=14, value=f"{total_net_profit/total_sales*100:.1f}%" if total_sales > 0 else "0%")
    
    # ============================================
    # СТРОКИ 7+: ДАННЫЕ ПО ДНЯМ
    # ============================================
    for idx, row_data in enumerate(data, start=7):
        date, day_of_week, sales, checks, cost_price, expenses = row_data
        
        avg_check = sales / checks if checks > 0 else 0
        gross_profit = sales - cost_price
        net_profit = gross_profit - expenses
        
        cost_percent = cost_price / sales * 100 if sales > 0 else 0
        gross_percent = gross_profit / sales * 100 if sales > 0 else 0
        expenses_percent = expenses / sales * 100 if sales > 0 else 0
        net_percent = net_profit / sales * 100 if sales > 0 else 0
        
        row = idx
        ws.cell(row=row, column=1, value=day_of_week).alignment = center_align
        ws.cell(row=row, column=2, value=date).alignment = center_align
        ws.cell(row=row, column=3, value=sales)
        ws.cell(row=row, column=4, value=checks)
        ws.cell(row=row, column=5, value=avg_check)
        ws.cell(row=row, column=6, value=cost_price)
        ws.cell(row=row, column=7, value=f"{cost_percent:.0f}%")
        ws.cell(row=row, column=8, value=cost_price)
        ws.cell(row=row, column=9, value=f"{cost_percent:.0f}%")
        ws.cell(row=row, column=10, value=gross_profit)
        ws.cell(row=row, column=11, value=f"{gross_percent:.0f}%")
        ws.cell(row=row, column=12, value=expenses)
        ws.cell(row=row, column=13, value=f"{expenses_percent:.0f}%")
        ws.cell(row=row, column=14, value=net_profit)
        ws.cell(row=row, column=15, value=f"{net_percent:.0f}%")
        
        # Цветовая индикация
        if net_profit > 0:
            ws.cell(row=row, column=14).fill = green_fill
        elif net_profit < 0:
            ws.cell(row=row, column=14).fill = red_fill
        
        if expenses_percent > 50:
            ws.cell(row=row, column=12).fill = yellow_fill
        
        # Границы
        for col in range(1, 16):
            ws.cell(row=row, column=col).border = thin_border
    
    # ============================================
    # ФОРМАТИРОВАНИЕ КОЛОНОК
    # ============================================
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 18
    ws.column_dimensions['O'].width = 10
    
    # Сохраняем файл
    filename = 'pnl_report.xlsx'
    wb.save(filename)
    
    return filename


def forecast_linear(x, y, steps=7):
    """
    Линейный прогноз на steps дней вперёд.
    Возвращает массив прогнозных значений.
    """
    if len(x) < 2:
        return np.full(steps, y[-1] if y else 0)
    
    # x - индексы (0,1,2,...), y - значения
    coeffs = np.polyfit(x, y, 1)
    poly = np.poly1d(coeffs)
    
    # Прогноз на следующие steps дней
    future_x = np.arange(len(x), len(x) + steps)
    forecast = poly(future_x)
    return forecast


def generate_dashboard():
    """
    Генерация дашборда в стиле Apple: воздушный, минималистичный, с KPI-карточками.
    """
    data = get_all_data()
    if not data:
        return None, None
    
    df = pd.DataFrame(data, columns=[
        'date', 'day', 'sales', 'checks', 'cost', 'expenses'
    ])
    
    # Расчёты
    df['gross_profit'] = df['sales'] - df['cost']
    df['net_profit'] = df['gross_profit'] - df['expenses']
    df['avg_check'] = df['sales'] / df['checks']
    
    # Прогноз на 7 дней (линейный тренд)
    x = np.arange(len(df))
    sales_forecast = forecast_linear(x, df['sales'].values, steps=7)
    profit_forecast = forecast_linear(x, df['net_profit'].values, steps=7)
    check_forecast = forecast_linear(x, df['avg_check'].values, steps=7)
    
    # Итоговые метрики
    total_sales = df['sales'].sum()
    total_profit = df['net_profit'].sum()
    total_expenses = df['expenses'].sum()
    total_cost = df['cost'].sum()
    avg_check = df['avg_check'].mean()
    days = len(df)
    
    # Тренды (наклон)
    if len(df) >= 2:
        sales_slope = np.polyfit(np.arange(len(df)), df['sales'], 1)[0]
        profit_slope = np.polyfit(np.arange(len(df)), df['net_profit'], 1)[0]
        check_slope = np.polyfit(np.arange(len(df)), df['avg_check'], 1)[0]
    else:
        sales_slope = profit_slope = check_slope = 0
    
    # ============================================
    # НАСТРОЙКА СТИЛЯ APPLE
    # ============================================
    plt.rcParams.update({
        'font.family': 'Helvetica Neue, Arial, sans-serif',
        'font.size': 11,
        'axes.edgecolor': '#E5E5EA',
        'axes.linewidth': 0.5,
        'axes.labelcolor': '#1C1C1E',
        'axes.titlecolor': '#1C1C1E',
        'xtick.color': '#8E8E93',
        'ytick.color': '#8E8E93',
        'grid.color': '#F2F2F7',
        'grid.linewidth': 0.5,
        'grid.alpha': 0.8,
        'figure.facecolor': '#F5F5F7',
        'axes.facecolor': '#FFFFFF',
    })
    
    # Цвета Apple
    APPLE_BLUE = '#007AFF'
    APPLE_GREEN = '#34C759'
    APPLE_RED = '#FF3B30'
    APPLE_ORANGE = '#FF9500'
    APPLE_GRAY = '#8E8E93'
    APPLE_DARK = '#1C1C1E'
    APPLE_LIGHT = '#F2F2F7'
    
    # ============================================
    # СОЗДАНИЕ ФИГУРЫ С GRIDSPEC
    # ============================================
    fig = plt.figure(figsize=(20, 14))
    
    # GridSpec: 3 строки (KPI, графики 2x2, пусто)
    gs = fig.add_gridspec(3, 4, height_ratios=[1.2, 3, 3], hspace=0.35, wspace=0.3)
    
    # ============================================
    # ЗАГОЛОВОК
    # ============================================
    fig.suptitle(
        'Аналитика ПНЛ',
        fontsize=28,
        fontweight='bold',
        color=APPLE_DARK,
        y=0.98,
        x=0.5
    )
    
    # Подзаголовок с периодом
    fig.text(
        0.5, 0.94,
        f'Период: {df["date"].iloc[0]} — {df["date"].iloc[-1]}  •  {days} дней',
        fontsize=13,
        color=APPLE_GRAY,
        ha='center'
    )
    
    # ============================================
    # KPI КАРТОЧКИ (верхняя строка)
    # ============================================
    kpi_data = [
        {
            'title': 'Продажи',
            'value': total_sales,
            'format': '{:,.0f} ₽',
            'color': APPLE_BLUE,
            'trend': sales_slope,
            'trend_label': 'в день'
        },
        {
            'title': 'Чистая прибыль',
            'value': total_profit,
            'format': '{:,.0f} ₽',
            'color': APPLE_GREEN if total_profit >= 0 else APPLE_RED,
            'trend': profit_slope,
            'trend_label': 'в день'
        },
        {
            'title': 'Расходы',
            'value': total_expenses,
            'format': '{:,.0f} ₽',
            'color': APPLE_RED,
            'trend': 0,
            'trend_label': ''
        },
        {
            'title': 'Средний чек',
            'value': avg_check,
            'format': '{:,.0f} ₽',
            'color': APPLE_ORANGE,
            'trend': check_slope,
            'trend_label': 'в день'
        }
    ]
    
    for i, kpi in enumerate(kpi_data):
        ax = fig.add_subplot(gs[0, i])
        ax.set_facecolor('#FFFFFF')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis('off')
        
        # Скруглённый прямоугольник (карточка)
        rect = plt.Rectangle(
            (0.02, 0.1), 0.96, 0.85,
            transform=ax.transAxes,
            facecolor='#FFFFFF',
            edgecolor='#E5E5EA',
            linewidth=1,
            clip_on=False,
            zorder=1
        )
        ax.add_patch(rect)
        
        # Цветная полоска сверху
        color_bar = plt.Rectangle(
            (0.02, 0.92), 0.96, 0.03,
            transform=ax.transAxes,
            facecolor=kpi['color'],
            clip_on=False,
            zorder=2
        )
        ax.add_patch(color_bar)
        
        # Заголовок карточки
        ax.text(
            0.5, 0.75, kpi['title'],
            transform=ax.transAxes,
            fontsize=12,
            color=APPLE_GRAY,
            ha='center',
            va='center',
            fontweight='normal'
        )
        
        # Значение
        ax.text(
            0.5, 0.50, kpi['format'].format(kpi['value']),
            transform=ax.transAxes,
            fontsize=22,
            color=APPLE_DARK,
            ha='center',
            va='center',
            fontweight='bold'
        )
        
        # Тренд
        if kpi['trend'] != 0:
            trend_icon = '↑' if kpi['trend'] > 0 else '↓'
            trend_color = APPLE_GREEN if kpi['trend'] > 0 else APPLE_RED
            ax.text(
                0.5, 0.25,
                f'{trend_icon} {abs(kpi["trend"]):,.0f} {kpi["trend_label"]}',
                transform=ax.transAxes,
                fontsize=11,
                color=trend_color,
                ha='center',
                va='center',
                fontweight='500'
            )
    
    # ============================================
    # ГРАФИК 1: ПРОДАЖИ И ПРОГНОЗ
    # ============================================
    ax1 = fig.add_subplot(gs[1, :2])
    
    # Факт — столбцы
    ax1.bar(
        range(len(df)), df['sales'],
        color=APPLE_BLUE, alpha=0.85,
        width=0.6, label='Факт', zorder=2
    )
    
    # Прогноз — пунктирная линия, продолжающая факт
    forecast_x = np.arange(len(df), len(df) + 7)
    ax1.plot(
        forecast_x, sales_forecast,
        '--', color=APPLE_ORANGE, linewidth=2.5,
        label='Прогноз (7 дней)', zorder=3, marker='o', markersize=4
    )
    
    # Вертикальная линия разделения факт/прогноз
    ax1.axvline(
        x=len(df) - 0.5,
        color=APPLE_GRAY, linewidth=1, linestyle=':', alpha=0.5, zorder=1
    )
    
    ax1.set_title('Продажи', fontsize=18, fontweight='bold', pad=15, loc='left')
    ax1.set_ylabel('Сумма, ₽', fontsize=12, color=APPLE_GRAY, labelpad=10)
    ax1.set_xticks(range(len(df) + 7))
    ax1.set_xticklabels(
        list(df['date']) + [f'+{i+1}' for i in range(7)],
        rotation=45, ha='right', fontsize=9, color=APPLE_GRAY
    )
    ax1.grid(axis='y', linestyle='-', alpha=0.5)
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.legend(loc='upper left', frameon=False, fontsize=10)
    
    # Аннотация с итогом
    ax1.text(
        0.98, 0.95, f'Итого: {total_sales:,.0f} ₽',
        transform=ax1.transAxes,
        fontsize=12, color=APPLE_DARK,
        ha='right', va='top',
        bbox=dict(facecolor='#FFFFFF', alpha=0.9, edgecolor='#E5E5EA', boxstyle='round,pad=0.5')
    )
    
    # ============================================
    # ГРАФИК 2: ЧИСТАЯ ПРИБЫЛЬ И ПРОГНОЗ
    # ============================================
    ax2 = fig.add_subplot(gs[1, 2:])
    
    colors_profit = [APPLE_GREEN if p >= 0 else APPLE_RED for p in df['net_profit']]
    ax2.bar(
        range(len(df)), df['net_profit'],
        color=colors_profit, alpha=0.85,
        width=0.6, label='Факт', zorder=2
    )
    
    ax2.plot(
        forecast_x, profit_forecast,
        '--', color=APPLE_ORANGE, linewidth=2.5,
        label='Прогноз (7 дней)', zorder=3, marker='o', markersize=4
    )
    
    ax2.axhline(y=0, color=APPLE_DARK, linewidth=0.8, linestyle='-', zorder=1)
    ax2.axvline(
        x=len(df) - 0.5,
        color=APPLE_GRAY, linewidth=1, linestyle=':', alpha=0.5, zorder=1
    )
    
    ax2.set_title('Чистая прибыль', fontsize=18, fontweight='bold', pad=15, loc='left')
    ax2.set_ylabel('Сумма, ₽', fontsize=12, color=APPLE_GRAY, labelpad=10)
    ax2.set_xticks(range(len(df) + 7))
    ax2.set_xticklabels(
        list(df['date']) + [f'+{i+1}' for i in range(7)],
        rotation=45, ha='right', fontsize=9, color=APPLE_GRAY
    )
    ax2.grid(axis='y', linestyle='-', alpha=0.5)
    ax2.spines['top'].set_visible(False)
    ax2.spines['right'].set_visible(False)
    ax2.legend(loc='upper left', frameon=False, fontsize=10)
    
    ax2.text(
        0.98, 0.95, f'Итого: {total_profit:,.0f} ₽',
        transform=ax2.transAxes,
        fontsize=12, color=APPLE_DARK,
        ha='right', va='top',
        bbox=dict(facecolor='#FFFFFF', alpha=0.9, edgecolor='#E5E5EA', boxstyle='round,pad=0.5')
    )
    
    # ============================================
    # ГРАФИК 3: СРЕДНИЙ ЧЕК И ПРОГНОЗ
    # ============================================
    ax3 = fig.add_subplot(gs[2, :2])
    
    ax3.plot(
        range(len(df)), df['avg_check'],
        '-', color=APPLE_BLUE, linewidth=2.5,
        label='Факт', zorder=3, marker='o', markersize=6, markerfacecolor='white', markeredgewidth=2
    )
    
    ax3.plot(
        forecast_x, check_forecast,
        '--', color=APPLE_ORANGE, linewidth=2.5,
        label='Прогноз (7 дней)', zorder=3, marker='o', markersize=4
    )
    
    ax3.axvline(
        x=len(df) - 0.5,
        color=APPLE_GRAY, linewidth=1, linestyle=':', alpha=0.5, zorder=1
    )
    
    ax3.set_title('Средний чек', fontsize=18, fontweight='bold', pad=15, loc='left')
    ax3.set_ylabel('Сумма, ₽', fontsize=12, color=APPLE_GRAY, labelpad=10)
    ax3.set_xticks(range(len(df) + 7))
    ax3.set_xticklabels(
        list(df['date']) + [f'+{i+1}' for i in range(7)],
        rotation=45, ha='right', fontsize=9, color=APPLE_GRAY
    )
    ax3.grid(axis='y', linestyle='-', alpha=0.5)
    ax3.spines['top'].set_visible(False)
    ax3.spines['right'].set_visible(False)
    ax3.legend(loc='upper left', frameon=False, fontsize=10)
    
    ax3.text(
        0.98, 0.95, f'Средний: {avg_check:,.0f} ₽',
        transform=ax3.transAxes,
        fontsize=12, color=APPLE_DARK,
        ha='right', va='top',
        bbox=dict(facecolor='#FFFFFF', alpha=0.9, edgecolor='#E5E5EA', boxstyle='round,pad=0.5')
    )
    
    # ============================================
    # ГРАФИК 4: СТРУКТУРА (ПИРОГ + МЕТРИКИ)
    # ============================================
    ax4 = fig.add_subplot(gs[2, 2:])
    
    # Данные для пирога
    if total_profit > 0:
        sizes = [total_cost, total_expenses, total_profit]
        labels = ['Себестоимость', 'Расходы', 'Прибыль']
        colors_pie = [APPLE_GRAY, APPLE_RED, APPLE_GREEN]
    else:
        sizes = [total_cost, total_expenses]
        labels = ['Себестоимость', 'Расходы']
        colors_pie = [APPLE_GRAY, APPLE_RED]
    
    wedges, texts, autotexts = ax4.pie(
        sizes, labels=None, colors=colors_pie,
        autopct='%1.1f%%', startangle=90,
        pctdistance=0.75,
        wedgeprops={'edgecolor': '#FFFFFF', 'linewidth': 2, 'width': 0.7}
    )
    
    # Стилизация процентов
    for autotext in autotexts:
        autotext.set_color('#FFFFFF')
        autotext.set_fontsize(13)
        autotext.set_fontweight('bold')
    
    ax4.set_title('Структура', fontsize=18, fontweight='bold', pad=15, loc='left')
    
    # Легенда справа с метриками
    legend_y = 0.85
    legend_x = 1.15
    for i, (label, size, color) in enumerate(zip(labels, sizes, colors_pie)):
        pct = size / sum(sizes) * 100
        ax4.text(
            legend_x, legend_y - i * 0.25,
            f'■ {label}',
            transform=ax4.transAxes,
            fontsize=12, color=APPLE_DARK,
            va='center'
        )
        ax4.text(
            legend_x + 0.35, legend_y - i * 0.25,
            f'{size:,.0f} ₽ ({pct:.1f}%)',
            transform=ax4.transAxes,
            fontsize=11, color=APPLE_GRAY,
            va='center', fontweight='500'
        )
    
    ax4.axis('equal')
    
    # ============================================
    # СОХРАНЕНИЕ
    # ============================================
    plt.tight_layout(rect=[0, 0, 1, 0.92])
    filename = 'dashboard.png'
    plt.savefig(
        filename, dpi=150, bbox_inches='tight',
        facecolor='#F5F5F7', edgecolor='none'
    )
    plt.close()
    
    # Генерируем текстовую аналитику
    analytics = generate_analytics_text(
        df, sales_forecast, profit_forecast, check_forecast,
        sales_slope, profit_slope, check_slope
    )
    
    return filename, analytics


def generate_analytics_text(df, sales_forecast, profit_forecast, check_forecast,
                           sales_slope, profit_slope, check_slope):
    """
    Генерация текстовой аналитики в стиле Apple — кратко, по делу, с эмодзи.
    """
    total_sales = df['sales'].sum()
    total_profit = df['net_profit'].sum()
    total_expenses = df['expenses'].sum()
    total_cost = df['cost'].sum()
    days = len(df)
    
    avg_daily_sales = total_sales / days if days > 0 else 0
    avg_daily_profit = total_profit / days if days > 0 else 0
    avg_check = df['avg_check'].mean() if days > 0 else 0
    
    # Прогнозные значения (суммарно за 7 дней)
    forecast_sales_sum = sales_forecast.sum() if len(sales_forecast) > 0 else 0
    forecast_profit_sum = profit_forecast.sum() if len(profit_forecast) > 0 else 0
    
    # Тренды
    sales_trend_icon = '📈' if sales_slope > 0 else '📉' if sales_slope < 0 else '➡️'
    profit_trend_icon = '📈' if profit_slope > 0 else '📉' if profit_slope < 0 else '➡️'
    
    # Анализ
    good = []
    bad = []
    
    if total_profit > 0:
        good.append(f"Прибыль положительная: {total_profit:,.0f} ₽")
    else:
        bad.append(f"Убыток: {total_profit:,.0f} ₽")
    
    if avg_daily_profit > 5000:
        good.append(f"Средний день прибыльный: {avg_daily_profit:,.0f} ₽")
    elif avg_daily_profit < 0:
        bad.append(f"Средний день убыточный: {avg_daily_profit:,.0f} ₽")
    
    if total_sales > 0:
        cost_pct = total_cost / total_sales * 100
        exp_pct = total_expenses / total_sales * 100
        
        if cost_pct < 40:
            good.append(f"Низкая себестоимость: {cost_pct:.1f}%")
        elif cost_pct > 50:
            bad.append(f"Высокая себестоимость: {cost_pct:.1f}%")
        
        if exp_pct < 40:
            good.append(f"Контролируемые расходы: {exp_pct:.1f}%")
        elif exp_pct > 60:
            bad.append(f"Расходы слишком высокие: {exp_pct:.1f}%")
    
    # Формирование текста
    text = "📊 АНАЛИТИКА ПНЛ\n\n"
    
    # Основные показатели
    text += "💰 ОСНОВНЫЕ ПОКАЗАТЕЛИ\n"
    text += f"• Продажи: {total_sales:,.0f} ₽\n"
    text += f"• Себестоимость: {total_cost:,.0f} ₽"
    if total_sales > 0:
        text += f" ({total_cost/total_sales*100:.1f}%)"
    text += "\n"
    text += f"• Расходы: {total_expenses:,.0f} ₽"
    if total_sales > 0:
        text += f" ({total_expenses/total_sales*100:.1f}%)"
    text += "\n"
    text += f"• Чистая прибыль: {total_profit:,.0f} ₽\n\n"
    
    # Средние
    text += "📈 СРЕДНИЕ ПОКАЗАТЕЛИ\n"
    text += f"• Продажи/день: {avg_daily_sales:,.0f} ₽\n"
    text += f"• Прибыль/день: {avg_daily_profit:,.0f} ₽\n"
    text += f"• Средний чек: {avg_check:,.0f} ₽\n\n"
    
    # Что хорошо
    if good:
        text += "✅ ХОРОШО\n"
        for item in good:
            text += f"• {item}\n"
        text += "\n"
    
    # Что плохо
    if bad:
        text += "⚠️ ТРЕБУЕТ ВНИМАНИЯ\n"
        for item in bad:
            text += f"• {item}\n"
        text += "\n"
    
    # Прогноз
    text += "🔮 ПРОГНОЗ НА 7 ДНЕЙ\n"
    text += f"{sales_trend_icon} Продажи: ~{forecast_sales_sum:,.0f} ₽"
    if sales_slope != 0:
        text += f" (тренд {sales_slope:+,.0f}/день)"
    text += "\n"
    text += f"{profit_trend_icon} Прибыль: ~{forecast_profit_sum:,.0f} ₽"
    if profit_slope != 0:
        text += f" (тренд {profit_slope:+,.0f}/день)"
    text += "\n"
    
    return text